from staffing_system import models
from django.shortcuts import render, redirect,HttpResponse

def depart_list(request):
    querySet = models.Department.objects.all()
    return render(request, 'departList.html', {"depart_list": querySet})


def depart_add(request):
    if request.method == "GET":
        return render(request, 'departAdd.html')
    if request.method == "POST":
        # 获取提交的表单数据
        title = request.POST.get("title")
        # 存储到数据库
        models.Department.objects.create(title=title)
        return redirect('/depart/list/')


def depart_delete(request):
    # 获取当前要删除的数据的id
    id = request.GET.get("id")
    # 数据库删除
    models.Department.objects.filter(id=id).delete()
    return redirect('/depart/list/')


def depart_edit(request, id):
    if request.method == "GET":
        obj = models.Department.objects.filter(id=id).first()
        return render(request, 'departEdit.html', {"obj": obj})
    if request.method == "POST":
        title = request.POST.get("title")
        models.Department.objects.filter(id=id).update(title=title)
        return redirect("/depart/list/")

# 上传excel文件，添加部门列表
def depart_multi(request):
    from openpyxl import load_workbook

    """批量上传文件（读取excel文件）"""
    # 获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    print(type(file_object))

    # 读取文件
    wb = load_workbook(file_object)
    # 读取第0个sheet
    sheet = wb.worksheets[0]
    # 读取第一行第二列
    # cell = sheet.cell(1,2)
    # print(cell.value)

    # 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        # row代表每一行
        # 获取某行的第一列数据
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect("/depart/list/")